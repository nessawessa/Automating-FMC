# -*- coding: utf-8 -*-

# This script creates Failure Modes based on data in the "Create FM's" worksheet in the FM&C Modification Template


import subprocess
import re
import json
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


def excel_read():

    # set up to surpress command window
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = subprocess.SW_HIDE

    wb_obj = openpyxl.load_workbook(r"C:\Users\aq34o\Documents\Automating FM&C\python scripts\FM&C Modification Template.xlsx")

    sheet=wb_obj['Start Here (Req\'d)']

    wwid = str((sheet.cell(row = 6, column = 3)).value)
    fmc_id = (sheet.cell(row = 11, column = 3)).value

    print("WWID = " + wwid)
    print("FM&C ID = " + str(fmc_id))
    print()

    # Determine the Library in which the FM&C resides

    imcmd = ("im issues --user=" + wwid + " --queryDefinition='((item.live)and"
         "(field[Type]=Failure Mode & Cause Document)and"
         "(field[\"ID\"]=" + str(fmc_id) + "))' "
         " --fields=Project")

    return_code = subprocess.check_output(imcmd, startupinfo=startupinfo)
    return_code = str(return_code, 'utf-8','ignore')
    fmc_export = return_code.splitlines()

    fmc_proj = fmc_export[0]
    print(fmc_proj)
    print()

    # Read Failure Mode Data

    sheet=wb_obj['Create Fail Modes']

    length = sheet.max_row
    row_count = length - 5
    row_num = 6

    fm_count = 0
    fail_mode_text = []

    while row_num <= row_count:

        fm_text = (sheet.cell(row = row_num, column = 2)).value
        #print(fm_text)
        if str(fm_text) != 'None':
            fail_mode_text.append(str(fm_text))
            fm_count += 1
        row_num += 1

    print("Failure Mode Count = " + str(fm_count))
    print()

    new_fm_ids = {}


    # Create new Failure Modes in RV&S

    row_num = 6
    for fm_txt in fail_mode_text:

        if len(str(fm_txt)) > 0:

            print(fm_txt)
            print(str(fmc_id))
            imcmd = ("im createcontent --hostname=integrity.cummins.com --port=7002 --user=" + wwid + " "
                 "--type=\"Failure Item\" --field=\"Category=Failure Mode\" --field=\"Type of Failure Item=Historical\" "
                 "--richContentfield=\"text=" + str(fm_txt) + " \" --ParentID=" + str(fmc_id) + " ")

            return_code = subprocess.run(imcmd, capture_output=True, text=True)
            output = return_code.stderr
            print(output)
            new_fm = re.findall(r'[0-9]+', output)
            fm_id = new_fm[0]
            print("New Failure Mode ID = " + str(fm_id))
            new_fm_ids.update({row_num: fm_id})

        row_num += 1

    print(new_fm_ids)

    # Write new Fail Mode ID's back to Excel Sheet

    row_num = 6
    row_start = row_num

    sheet=wb_obj['Create Fail Modes']

    while row_num <= row_start + int(fm_count):

        fm_id = new_fm_ids.get(row_num)
        print(fm_id)
        sheet.cell(row=row_num, column=4).value = fm_id
        row_num += 1

    wb_obj.save(r"C:\Users\aq34o\Documents\Automating FM&C\python scripts\FM&C Modification Template.xlsx")

    flag = 'FM&C updates completed successfully!'
    return flag

############################################################################################################

# main program

current_dir = os.getcwd()
excel_file = "FM&C Modification Template.xlsx"
input_file_path = os.path.join(current_dir, excel_file)

print()
print('Create Failure Modes based on Excel data')
print()

FLAG = excel_read()

print()
print(FLAG)
