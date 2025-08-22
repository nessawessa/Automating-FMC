import openpyxl
import subprocess
import sys
import re

EXCEL_PATH = r"C:\Users\aq34o\Documents\Automating FM&C\python scripts\FM&C Modification Template.xlsx"
SHEET_REVISE = "Revise Causes"
SHEET_START = "Start Here (Req'd)"
HOSTNAME = "integrity.cummins.com"
PORT = "7002"

# Hard mappings from Excel to RV&S field names
FIELD_DETECTION_PRIMARY = "Library Detection Rating"
FIELD_PREVENTION_PRIMARY = "Library Prevention Rating"
FIELD_DAMAGE_CATEGORY = "Damage Category"
# Damage Mechanism was not updating; keep a resolver + candidates
MECHANISM_CANDIDATES = [
    "Damage Mechanism", "Damage Mechanisms", "Library Damage Mechanism",
    "Failure Mechanism", "Mechanism of Damage", "Degradation Mechanism", "Mechanism"
]

RICH_TEXT_LABEL = "Text"
RICH_TRANSFER_LABEL = "Transfer Function"

def _hide_window_startupinfo():
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    si.wShowWindow = subprocess.SW_HIDE
    return si

def _run_cmd(cmd: str) -> tuple[int, str]:
    proc = subprocess.run(
        cmd, capture_output=True, text=True, shell=True, startupinfo=_hide_window_startupinfo()
    )
    out = ""
    if proc.stdout:
        out += proc.stdout
    if proc.stderr:
        out += ("\n" if out else "") + proc.stderr
    return proc.returncode, out

def _escape_rich_text(val) -> str:
    s = "" if val is None else str(val)
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    s = s.replace('"', '\\"')
    return s

def _escape_plain(val) -> str:
    if val is None:
        return ""
    return str(val).strip().replace('"', '\\"')

def _set_field(connect: str, issue_id: str, field_name: str, value: str) -> tuple[int, str]:
    # Try overwrite first
    cmd_set = f'im editissue{connect} --field="{field_name}={value}" {issue_id}'
    rc, out = _run_cmd(cmd_set)
    if rc == 0:
        return rc, out
    # If looks like list-type or overwrite rejected, try addFieldValues
    if re.search(r'list|multi-?valued|use\s*addFieldValues', out, re.IGNORECASE):
        cmd_add = f'im editissue{connect} --addFieldValues="{field_name}={value}" {issue_id}'
        return _run_cmd(cmd_add)
    # Generic fallback once
    cmd_add2 = f'im editissue{connect} --addFieldValues="{field_name}={value}" {issue_id}'
    rc2, out2 = _run_cmd(cmd_add2)
    if rc2 == 0:
        return rc2, out2
    return rc, out

def _field_exists(connect: str, issue_id: str, field_name: str) -> bool:
    q = f'(field[\\"ID\\"]={issue_id})'
    cmd = f'im issues{connect} --queryDefinition="{q}" --fields="{field_name}"'
    rc, out = _run_cmd(cmd)
    if rc != 0:
        return False
    if re.search(r'does not exist|unknown field|invalid field', out, re.IGNORECASE):
        return False
    return True

def _list_field_names(connect: str, issue_id: str) -> list[str]:
    cmd = f'im viewissue{connect} {issue_id} --showFields'
    rc, out = _run_cmd(cmd)
    if rc != 0:
        return []
    fields = []
    for line in out.splitlines():
        if ":" in line:
            label = line.split(":", 1)[0].strip()
            if label and len(label) < 100:
                fields.append(label)
    # de-dup preserving order
    seen = set()
    uniq = []
    for f in fields:
        if f not in seen:
            uniq.append(f); seen.add(f)
    return uniq

def _resolve_mechanism_field(connect: str, issue_id: str) -> str | None:
    discovered = _list_field_names(connect, issue_id)
    # Prefer any field that contains "mechan" in its label
    for f in discovered:
        if "mechan" in f.lower():
            return f
    # Fall back to probing candidate names
    for cand in MECHANISM_CANDIDATES:
        if _field_exists(connect, issue_id, cand):
            return cand
    return None

def main():
    # Load workbook (we do NOT write to Column I; preserve your CONCATENATE formulas)
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except Exception as e:
        print(f"Failed to open workbook: {e}")
        sys.exit(1)

    # Read WWID from Start sheet C6
    try:
        start_sheet = wb[SHEET_START]
        user = str(start_sheet.cell(row=6, column=3).value or "").strip()
        if not user:
            raise ValueError("Missing WWID in 'Start Here (Req'd)' at C6.")
    except Exception as e:
        print(f"Failed to read user WWID: {e}")
        sys.exit(1)

    try:
        sheet = wb[SHEET_REVISE]
    except KeyError:
        print(f"Worksheet '{SHEET_REVISE}' not found.")
        sys.exit(1)

    connect = f" --hostname={HOSTNAME} --port={PORT} --user={user}"

    row = 6
    processed = 0
    while True:
        cause_id = sheet.cell(row=row, column=2).value   # B: Cause ID
        cause_text = sheet.cell(row=row, column=3).value # C: Cause Text
        prevention = sheet.cell(row=row, column=4).value # D: Prevention (Excel) -> Library Prevention Rating (RV&S)
        detection = sheet.cell(row=row, column=5).value  # E: Detection (Excel) -> Library Detection Rating (RV&S)
        damage_cat = sheet.cell(row=row, column=6).value # F: Damage Category
        damage_mech = sheet.cell(row=row, column=7).value# G: Damage Mechanism
        xfer_fn = sheet.cell(row=row, column=8).value    # H: Transfer Function
        # I: Revise_Script (formula) — DO NOT MODIFY
        # J: Failure Mode (informational)

        if not cause_id and not cause_text:
            break

        if cause_id:
            cid = str(cause_id).strip()
            print(f"[Row {row}] Editing Cause ID {cid}")

            # 1) Text
            if cause_text is not None and str(cause_text) != "":
                val = _escape_rich_text(cause_text)
                cmd = f'im editissue{connect} --RichContentField="{RICH_TEXT_LABEL}={val}" {cid}'
                rc, out = _run_cmd(cmd)
                print(f"[Row {row}] Text rc={rc}")
                if out.strip():
                    print(f"[Row {row}] Text output:\n{out}")

            # 2) Detection -> Library Detection Rating
            if detection is not None and str(detection).strip() != "":
                value = _escape_plain(detection)
                field_used = None
                # Try primary mapped name
                if _field_exists(connect, cid, FIELD_DETECTION_PRIMARY):
                    field_used = FIELD_DETECTION_PRIMARY
                else:
                    # Common fallbacks just in case
                    for alt in ["Detection", "Detection Rating", "ESW Detection Rating"]:
                        if _field_exists(connect, cid, alt):
                            field_used = alt
                            break
                if field_used:
                    rc, out = _set_field(connect, cid, field_used, value)
                    print(f"[Row {row}] Detection rc={rc} (field='{field_used}')")
                    if out.strip():
                        print(f"[Row {row}] Detection output:\n{out}")
                else:
                    print(f"[Row {row}] Detection-like field not found on item; skipping Detection.")

            # 3) Prevention -> Library Prevention Rating
            if prevention is not None and str(prevention).strip() != "":
                value = _escape_plain(prevention)
                field_used = None
                if _field_exists(connect, cid, FIELD_PREVENTION_PRIMARY):
                    field_used = FIELD_PREVENTION_PRIMARY
                else:
                    for alt in ["Prevention", "Prevention Rating", "ESW Prevention Rating"]:
                        if _field_exists(connect, cid, alt):
                            field_used = alt
                            break
                if field_used:
                    rc, out = _set_field(connect, cid, field_used, value)
                    print(f"[Row {row}] Prevention rc={rc} (field='{field_used}')")
                    if out.strip():
                        print(f"[Row {row}] Prevention output:\n{out}")
                else:
                    print(f"[Row {row}] Prevention-like field not found on item; skipping Prevention.")

            # 4) Damage Category (worked previously)
            if damage_cat is not None and str(damage_cat).strip() != "":
                rc, out = _set_field(connect, cid, FIELD_DAMAGE_CATEGORY, _escape_plain(damage_cat))
                print(f"[Row {row}] Damage Category rc={rc} (field='{FIELD_DAMAGE_CATEGORY}')")
                if out.strip():
                    print(f"[Row {row}] Damage Category output:\n{out}")

            # 5) Damage Mechanism (resolve dynamically until we know the exact label)
            if damage_mech is not None and str(damage_mech).strip() != "":
                mech_field = _resolve_mechanism_field(connect, cid)
                if mech_field:
                    rc, out = _set_field(connect, cid, mech_field, _escape_plain(damage_mech))
                    print(f"[Row {row}] Damage Mechanism rc={rc} (field='{mech_field}')")
                    if out.strip():
                        print(f"[Row {row}] Damage Mechanism output:\n{out}")
                else:
                    print(f"[Row {row}] Could not resolve a Damage Mechanism field on this item; skipping.")

            # 6) Transfer Function
            if xfer_fn is not None and str(xfer_fn) != "":
                val = _escape_rich_text(xfer_fn)
                cmd = f'im editissue{connect} --RichContentField="{RICH_TRANSFER_LABEL}={val}" {cid}'
                rc, out = _run_cmd(cmd)
                print(f"[Row {row}] Transfer Function rc={rc}")
                if out.strip():
                    print(f"[Row {row}] Transfer Function output:\n{out}")

            processed += 1

        row += 1

    # Do not save workbook: we didn't modify cells; Column I formulas remain intact
    print(f"Processed rows: {processed}. Revisions executed. Column I formulas preserved.")

if __name__ == "__main__":
    main()