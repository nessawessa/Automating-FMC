import openpyxl
import subprocess
import re
import os
import sys

EXCEL_PATH = r"C:\Users\aq34o\Documents\Automating FM&C\python scripts\FM&C Modification Template.xlsx"
SHEET_CAUSES = "Create Causes"
SHEET_START = "Start Here (Req'd)"
HOSTNAME = "integrity.cummins.com"
PORT = "7002"

def _hide_window_startupinfo():
    si = subprocess.STARTUPINFO()
    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    si.wShowWindow = subprocess.SW_HIDE
    return si

def _run_cmd(cmd: str) -> tuple[int, str]:
    startupinfo = _hide_window_startupinfo()
    proc = subprocess.run(cmd, capture_output=True, text=True, shell=True, startupinfo=startupinfo)
    out = ""
    # Combine in a consistent order for parsing
    if proc.stdout:
        out += proc.stdout
    if proc.stderr:
        out += ("\n" if out else "") + proc.stderr
    return proc.returncode, out

def _escape_rich_text(text: str) -> str:
    # Escape double quotes for RichContentField
    return (text or "").replace('"', '\\"')

def _parse_created_id(output: str, exclude_ids=None) -> str | None:
    exclude_ids = set(map(str, exclude_ids or []))
    patterns = [
        r'Created\s+(?:item|content)\s+(\d+)',     # e.g., "Created item 12345678"
        r'created\s+content.*?\bID[:\s]+(\d+)',    # variants
        r'\bID[:\s]+(\d+)\b',                      # generic "ID: 12345678"
    ]
    for p in patterns:
        m = re.search(p, output, re.IGNORECASE | re.DOTALL)
        if m and m.group(1) not in exclude_ids:
            return m.group(1)
    # Fallback: pick a long-ish number not excluded
    for n in re.findall(r'\b\d{6,}\b', output):
        if n not in exclude_ids:
            return n
    return None

def create_causes():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except Exception as e:
        print(f"Failed to open workbook: {e}")
        sys.exit(1)

    # Read WWID/user from Start sheet (C6)
    try:
        start_sheet = wb[SHEET_START]
        user = str(start_sheet.cell(row=6, column=3).value or "").strip()
        if not user:
            raise ValueError("Missing WWID in 'Start Here (Req'd)'! Expected cell C6.")
    except Exception as e:
        print(f"Failed to read user WWID: {e}")
        sys.exit(1)

    try:
        sheet = wb[SHEET_CAUSES]
    except KeyError:
        print(f"Worksheet '{SHEET_CAUSES}' not found.")
        sys.exit(1)

    connect = f" --hostname={HOSTNAME} --port={PORT} --user={user}"

    row = 6
    processed = 0
    while True:
        fm_id = sheet.cell(row=row, column=2).value  # Column B: Fail Mode ID
        cause_text = sheet.cell(row=row, column=3).value  # Column C: Cause Text

        # Stop at first empty pair (adjust if you want to skip blanks instead)
        if not fm_id and not cause_text:
            break

        if fm_id and cause_text:
            fm_id_str = str(fm_id).strip()
            text_escaped = _escape_rich_text(str(cause_text))

            # IMPORTANT: Do not insert "after:<parentID>" – use default/last to avoid MKS124707
            create_script = (
                f'im createcontent{connect} --type="Failure Item" '
                f'--field="Category=Failure Cause" '
                f'--field="Type of Failure Item=Historical" '
                f'--RichContentField="Text={text_escaped}" '
                f'--parentID={fm_id_str} --insertLocation=last'
            )
            sheet.cell(row=row, column=4).value = create_script  # Column D

            rc, out = _run_cmd(create_script)
            print(f"[Row {row}] Create rc={rc}")
            if out.strip():
                print(f"[Row {row}] Create output:\n{out}")

            cause_id = None
            if rc == 0:
                cause_id = _parse_created_id(out, exclude_ids=[fm_id_str, PORT])
            else:
                # If the only issue was an invalid insertLocation, retry once without it
                if "MKS124822" in out or "insertLocation" in out:
                    retry_script = (
                        f'im createcontent{connect} --type="Failure Item" '
                        f'--field="Category=Failure Cause" '
                        f'--field="Type of Failure Item=Historical" '
                        f'--RichContentField="Text={text_escaped}" '
                        f'--parentID={fm_id_str}'
                    )
                    print(f"[Row {row}] Retrying create without insertLocation...")
                    rc, out = _run_cmd(retry_script)
                    print(f"[Row {row}] Retry rc={rc}")
                    if out.strip():
                        print(f"[Row {row}] Retry output:\n{out}")
                    if rc == 0:
                        cause_id = _parse_created_id(out, exclude_ids=[fm_id_str, PORT])

            if not cause_id:
                print(f"[Row {row}] Unable to determine created Cause ID. Skipping relate.")
                sheet.cell(row=row, column=5).value = ""  # Created Cause ID (E)
                sheet.cell(row=row, column=6).value = ""  # Relate_Script (F)
            else:
                # Write created Cause ID
                sheet.cell(row=row, column=5).value = cause_id

                relate_script = (
                    f'im editissue{connect} --field="Failure Cause to Mode={fm_id_str}" {cause_id}'
                )
                sheet.cell(row=row, column=6).value = relate_script  # Column F

                rc2, out2 = _run_cmd(relate_script)
                print(f"[Row {row}] Relate rc={rc2}")
                if out2.strip():
                    print(f"[Row {row}] Relate output:\n{out2}")
                if rc2 != 0:
                    print(f"[Row {row}] Relate failed. Please review the output above.")

            processed += 1

        row += 1

    # Save updates to Excel
    try:
        wb.save(EXCEL_PATH)
        print(f"Processed rows: {processed}. Causes created/related and Excel updated successfully!")
    except PermissionError:
        print("PermissionError: Close the Excel workbook before running this script and try again.")
    except Exception as e:
        print(f"Failed to save workbook: {e}")

if __name__ == "__main__":
    create_causes()