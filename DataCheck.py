# -*- coding: utf-8 -*-

# This script verifies the structure of the Conversational Question-style data loaded in Question_Upload_Template.xlsx
# It is intended to be run prior to QuestionUpload.py to confirm the input data structure before
# any new Change Configurations are created since Change Configurations cannot be (easily) deleted


import subprocess
import re
import json
import openpyxl
import requests
from openpyxl.styles import Font
import os


def excel_read():

    # set up to surpress command window
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = subprocess.SW_HIDE

    wb_obj = openpyxl.load_workbook(input_file_path)

    # Variable row_offset is the number of rows above the start of the table - which starts with the Cause ID's
    row_offset = 4
    # Variable col_offset is the number of columns to the left of the table - which starts with the Level 1 Questions
    col_offset = 1

    sheet=wb_obj['CC_Questions']

    wwid = str((sheet.cell(row = 2, column = 4)).value)
    fmc_id = (sheet.cell(row = 3, column = 4)).value

    print("WWID = " + wwid)
    print("FM&C ID = " + str(fmc_id))
    print()

    row_count = sheet.max_row
    quest_count = row_count - (row_offset + 1)
    col_count = sheet.max_column
    cause_count = col_count - (col_offset + 2)

    print("Question Count =" + str(quest_count))
    print("Cause Count =" + str(cause_count))
    print()

    cause_ids = []
    quest_seqs = []
    questions = {}
    quest2cause_map = {}
    seq_pairs = {}

    # Read and store Cause ID's

    row_num = row_offset + 1
    col_num = col_offset + 3

    while col_num <= col_count:

        id = (sheet.cell(row = row_num, column = col_num)).value
        cause_ids.append(id)
        col_num += 1

    print("Cause ID's:")
    print(cause_ids)
    print()

    # Read and store Questions

    L1_num = 1
    L2_num = 1
    row_num = row_offset + 2
    while row_num <= row_count:

        L1_quest = str((sheet.cell(row = row_num, column = (col_offset + 1))).value)
        L2_quest = str((sheet.cell(row = row_num, column = (col_offset + 2))).value)

        if L1_quest != "None" and L2_quest == "None":

            this_seq = str(L1_num)
            quest_seqs.append(this_seq)
            questions.update({this_seq: L1_quest})
            L1_curr = L1_num
            L2_num = 1
            L1_num += 1

        elif L1_quest == "None" and L2_quest != "None":

            this_seq = str(L1_curr) + "-" + str(L2_num)
            quest_seqs.append(this_seq)
            questions.update({this_seq: L2_quest})
            L2_num += 1

        else:

            print("Error Detected in Question Data!")

        row_num += 1

    print("Sequence Identifiers:")
    print(quest_seqs)
    print()
    print("Question Data:")
    print(questions)
    print()


    # Read and store Question to Cause Mapping

    row_num = row_offset + 2

    while row_num <= row_count:

        cause_list = []
        row_idx = row_num - (row_offset + 2)
        col_num = col_offset + 3
        this_seq = quest_seqs[row_idx]

        while col_num <= col_count:

            cause_idx = col_num - 4
            this_id = cause_ids[cause_idx]

            flag = str((sheet.cell(row = row_num, column = col_num)).value)

            if flag == "x":

                cause_list.append(this_id)
            
            col_num += 1

        quest2cause_map.update({this_seq: cause_list})
        row_num += 1

    print("Cause Mapping Data:")
    print(quest2cause_map)
    print()

    # Confirm that Causes exist

    for cause in cause_ids:

        imcmd = ("im issues --user=" + wwid + " --queryDefinition='((item.live)and"
            "(field[Type]=Failure Item)and"
            "(field[Category]=Failure Cause)and"
            "(field[\"Document ID Copy\"]=" + str(fmc_id) + ")and"
            "(field[\"ID\"]=" + str(cause) + "))' "
            " --fields=Summary")

        return_code = subprocess.check_output(imcmd, startupinfo=startupinfo)

    print("All Cause ID's have been verified!")

    return wwid,fmc_id,cause_ids,quest_seqs,questions,quest2cause_map


# main program

current_dir = os.getcwd()
excel_file = "Question_Upload_Template.xlsx"
input_file_path = os.path.join(current_dir, excel_file)

print()
print('Read Question & Cause Mapping Data from Excel')
print()

WWID,FMCID,CAUSEIDS,QUESTSEQS,QUESTIONS,QUEST2CAUSE = excel_read()

print("Data input completed successfully")
print()
print("Review Sequence & Question Data above to confirm accuracy")

